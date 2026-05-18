"""
report_generator.py — Professional PDF + Excel Reports
=======================================================
PDF Layout (4–5 pages):
  Page 1 : Executive Dashboard  — KPI cards + AI summary
  Page 2 : Visual Analytics     — 6 polished charts in 2-col grid
  Page 3 : Top Items table      + AI insights (two-column layout)
  Page 4 : Dataset Sample       — styled first-20-rows preview
  Page 5 : Forecast             — only rendered when forecast data exists

ML insights removed entirely.
Charts upgraded: richer colours, annotations, gradients, spines, custom ticks.
Excel upgraded: conditional formatting, sparkline-style bars, richer charts.
"""

import io, datetime, warnings
import numpy as np
import pandas as pd
warnings.filterwarnings("ignore")

# ── matplotlib ────────────────────────────────────────────────────────────────
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import matplotlib.patheffects as pe
    from matplotlib.patches import FancyBboxPatch
    import matplotlib.colors as mcolors
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

# ── ReportLab ─────────────────────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        Image as RLImage, HRFlowable, PageBreak,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAS_RL = True
except ImportError:
    HAS_RL = False

# ── OpenPyXL ──────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    HAS_OXL = True
except ImportError:
    HAS_OXL = False

NOW = datetime.datetime.now()

# ════════════════════════════════════════════════════════════════════════════════
# DESIGN SYSTEM
# ════════════════════════════════════════════════════════════════════════════════
BRAND = {
    "deep":    "#4C1D95",
    "mid":     "#7C3AED",
    "soft":    "#A78BFA",
    "pale":    "#EDE9FE",
    "teal":    "#0D9488",
    "teal_lt": "#CCFBF1",
    "ink":     "#1E1B4B",
    "muted":   "#6B7280",
    "rule":    "#DDD6FE",
    "white":   "#FFFFFF",
    "offwhite":"#FAFAF9",
    "green":   "#16A34A",
    "red":     "#DC2626",
    "row_a":   "#F5F3FF",
    "row_b":   "#FFFFFF",
}

# Chart palettes — each chart gets its own cohesive multi-stop palette
CHART_PALETTES = {
    "region":   ["#0F766E","#0D9488","#14B8A6","#2DD4BF","#5EEAD4","#99F6E4"],
    "category": ["#4C1D95","#6D28D9","#7C3AED","#8B5CF6","#A78BFA","#C4B5FD","#DDD6FE"],
    "trend":    "#059669",
    "scatter":  "#EA580C",
    "orders":   "#2563EB",
    "heatmap":  "BuPu",
}

KPI_CARDS = [
    ("#ECFDF5", "#059669"),
    ("#EFF6FF", "#2563EB"),
    ("#F5F3FF", "#7C3AED"),
    ("#FFF7ED", "#EA580C"),
    ("#F0FDF4", "#16A34A"),
    ("#FEF2F2", "#DC2626"),
]

XL_ACCENTS = {
    "README":          "4C1D95",
    "KPIs":            "0D9488",
    "Sales_by_Region": "0369A1",
    "Sales_by_Category":"7C3AED",
    "Top_Products":    "B45309",
    "Monthly_Trend":   "047857",
    "Raw_Data":        "374151",
}

# ════════════════════════════════════════════════════════════════════════════════
# DATA HELPERS
# ════════════════════════════════════════════════════════════════════════════════
def _col_types(df):
    return {
        "numeric":     df.select_dtypes(include="number").columns.tolist(),
        "categorical": df.select_dtypes(include=["object","category"]).columns.tolist(),
        "datetime":    df.select_dtypes(include=["datetime64","datetimetz"]).columns.tolist(),
    }

def _sales_col(df, ct):
    for kw in ["total revenue","revenue","sales","amount","total"]:
        c = next((c for c in ct["numeric"] if kw in c.lower()), None)
        if c: return c
    return ct["numeric"][0] if ct["numeric"] else None

def _profit_col(df, ct):
    return next((c for c in ct["numeric"]
                 if "profit" in c.lower() and "margin" not in c.lower()), None)

def _cat_col(df, ct):
    for kw in ["categ","type","segment","cuisine","food_type"]:
        c = next((c for c in ct["categorical"] if kw in c.lower()), None)
        if c: return c
    for c in ct["categorical"]:
        if df[c].nunique() <= 30 and not any(x in c.lower() for x in ["id","name","order","customer"]):
            return c
    return None

def _item_col(df, ct):
    for exact in ["product name","food item","item name","sku name"]:
        c = next((c for c in df.columns if c.lower() == exact), None)
        if c: return c
    for kw in ["product","food","item","sku"]:
        c = next((c for c in ct["categorical"] if kw in c.lower()), None)
        if c: return c
    return None

def _reg_col(df, ct):
    return next((c for c in df.columns if "region" in c.lower()), None)

def _date_col(df, ct):
    dcs = ct.get("datetime", [])
    if dcs: return dcs[0]
    for c in df.columns:
        if "date" in c.lower(): return c
    return None

def _fmt(v):
    try:
        v = float(v)
        if abs(v) >= 1e9: return f"${v/1e9:.2f}B"
        if abs(v) >= 1e6: return f"${v/1e6:.2f}M"
        if abs(v) >= 1e3: return f"${v/1e3:.1f}K"
        return f"${v:,.0f}"
    except: return str(v)

def _num(x):
    try: return float(x)
    except: return 0.0

def _sales_by_region(df, ct):
    rc, sc = _reg_col(df, ct), _sales_col(df, ct)
    if not rc or not sc: return pd.DataFrame({"Region": ["N/A"], "Sales": [0]})
    out = df.groupby(rc)[sc].sum().reset_index().sort_values(sc, ascending=True)
    out.columns = ["Region", "Sales"]
    return out.round(2)

def _sales_by_category(df, ct):
    cc, sc = _cat_col(df, ct), _sales_col(df, ct)
    if not cc or not sc: return pd.DataFrame({"Category": ["N/A"], "Sales": [0], "Pct": [100]})
    out = df.groupby(cc)[sc].sum().reset_index().sort_values(sc, ascending=True)
    out.columns = ["Category", "Sales"]
    tot = out["Sales"].sum()
    out["Pct"] = (out["Sales"] / tot * 100).round(1) if tot else 0
    return out.round(2)

def _top_items(df, ct, n=16):
    item = _item_col(df, ct) or _cat_col(df, ct)
    sc, pc = _sales_col(df, ct), _profit_col(df, ct)
    if not item or not sc: return pd.DataFrame({"Item": ["N/A"], "Sales": [0]})
    gcols = [sc] + ([pc] if pc else [])
    out = (df.groupby(item)[gcols].sum().reset_index()
             .sort_values(sc, ascending=False).head(n))
    if pc:
        tot = out[sc].sum()
        out["Profit Rank %"] = (out[sc] / tot * 100).round(1)
    out.insert(0, "Rank", range(1, len(out)+1))
    cols = ["Rank", item, sc] + ([pc, "Profit Rank %"] if pc else [])
    out = out[cols]
    out.columns = ["Rank", "Item", "Sales"] + (["Profit", "Profit Rank %"] if pc else [])
    return out.round(2)

def _monthly_trend(df, ct):
    dc, sc = _date_col(df, ct), _sales_col(df, ct)
    if not dc or not sc: return pd.DataFrame({"Date": [], "Sales": [], "MoM": []})
    tmp = df.copy()
    tmp[dc] = pd.to_datetime(tmp[dc], errors="coerce")
    tmp["_p"] = tmp[dc].dt.to_period("M").dt.to_timestamp()
    m = tmp.groupby("_p")[sc].sum().reset_index()
    m.columns = ["Date", "Sales"]
    m["MoM"] = m["Sales"].pct_change().fillna(0).round(4)
    return m.round(2)


# ════════════════════════════════════════════════════════════════════════════════
# CHART FUNCTIONS — upgraded visuals
# ════════════════════════════════════════════════════════════════════════════════
def _save(fig, dpi=180):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight",
                facecolor=fig.get_facecolor(), edgecolor="none")
    buf.seek(0)
    plt.close(fig)
    return buf

def _base_fig(w=5.6, h=3.0):
    fig, ax = plt.subplots(figsize=(w, h))
    fig.patch.set_facecolor("#FAFAFA")
    ax.set_facecolor("#FFFFFF")
    return fig, ax

def _clean_ax(ax, grid_axis="x"):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#E5E7EB")
    ax.spines["bottom"].set_color("#E5E7EB")
    ax.tick_params(colors="#6B7280", labelsize=7.5, length=3)
    ax.set_axisbelow(True)
    if grid_axis == "x":
        ax.xaxis.grid(True, color="#F0F0F0", linewidth=0.7, linestyle="--")
        ax.yaxis.grid(False)
    else:
        ax.yaxis.grid(True, color="#F0F0F0", linewidth=0.7, linestyle="--")
        ax.xaxis.grid(False)

def _chart_title(ax, text, color, subtitle=None):
    ax.set_title(text, fontsize=10, fontweight="bold", color=color,
                 pad=10, loc="left", fontfamily="sans-serif")
    if subtitle:
        ax.text(0, 1.01, subtitle, transform=ax.transAxes,
                fontsize=7, color="#9CA3AF", va="bottom")


# ── 1. Sales by Region — gradient horizontal bars ────────────────────────────
def chart_region(df, ct):
    data = _sales_by_region(df, ct)
    if data.empty or data["Sales"].sum() == 0: return None
    sc_lbl = _sales_col(df, ct) or "Sales"
    n = len(data)
    palette = CHART_PALETTES["region"]
    clrs = [palette[min(i, len(palette)-1)] for i in range(n)]

    fig, ax = _base_fig(5.6, max(2.6, n * 0.42 + 0.6))
    bars = ax.barh(data["Region"].astype(str), data["Sales"],
                   color=clrs, edgecolor="white", linewidth=0.6,
                   height=0.6, zorder=3)
    max_v = data["Sales"].max()
    for bar, clr in zip(bars, clrs):
        w = bar.get_width()
        # value label inside bar (right-aligned white)
        ax.text(w - max_v * 0.015, bar.get_y() + bar.get_height()/2,
                _fmt(w), va="center", ha="right", fontsize=7.5,
                color="white", fontweight="bold",
                path_effects=[pe.withStroke(linewidth=1.5, foreground=clr)])
    # Subtle % of total annotation outside
    for bar in bars:
        w = bar.get_width()
        pct = w / data["Sales"].sum() * 100
        ax.text(max_v * 1.02, bar.get_y() + bar.get_height()/2,
                f"{pct:.1f}%", va="center", ha="left", fontsize=6.5, color="#9CA3AF")

    ax.set_xlim(0, max_v * 1.18)
    _chart_title(ax, f"{sc_lbl} by Region", "#0F766E",
                 f"Total: {_fmt(data['Sales'].sum())}")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: _fmt(x)))
    _clean_ax(ax, "x")
    fig.tight_layout(pad=0.7)
    return _save(fig)


# ── 2. Sales by Category — lollipop chart ───────────────────────────────────
def chart_category(df, ct):
    data = _sales_by_category(df, ct)
    if data.empty or data["Sales"].sum() == 0: return None
    sc_lbl = _sales_col(df, ct) or "Sales"
    cat_lbl = _cat_col(df, ct) or "Category"
    # Sort descending for lollipop
    data = data.sort_values("Sales", ascending=True)
    n = len(data)
    palette = CHART_PALETTES["category"]
    clrs = [palette[min(i, len(palette)-1)] for i in range(n)]

    fig, ax = _base_fig(5.6, max(2.8, n * 0.44 + 0.6))
    y_pos = range(n)
    # Stem lines
    for i, (y, val, clr) in enumerate(zip(y_pos, data["Sales"], clrs)):
        ax.plot([0, val], [y, y], color=clr, linewidth=1.8, alpha=0.4, zorder=2)
    # Dots
    ax.scatter(data["Sales"], y_pos, color=clrs, s=80, zorder=4,
               edgecolors="white", linewidths=1.2)
    # Labels
    max_v = data["Sales"].max()
    for y, val, pct, clr in zip(y_pos, data["Sales"], data["Pct"], clrs):
        ax.text(val + max_v * 0.02, y, f"{_fmt(val)}  ({pct:.1f}%)",
                va="center", fontsize=6.8, color="#374151")

    ax.set_yticks(list(y_pos))
    ax.set_yticklabels(data["Category"].astype(str), fontsize=7.5)
    ax.set_xlim(0, max_v * 1.35)
    _chart_title(ax, f"{sc_lbl} by {cat_lbl}", "#6D28D9")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: _fmt(x)))
    _clean_ax(ax, "x")
    ax.yaxis.grid(False)
    fig.tight_layout(pad=0.7)
    return _save(fig)


# ── 3. Monthly trend — area + line + peak annotation ────────────────────────
def chart_trend(df, ct):
    mt = _monthly_trend(df, ct)
    if mt.empty: return None
    sc_lbl = _sales_col(df, ct) or "Sales"
    color = CHART_PALETTES["trend"]

    fig, ax = _base_fig(5.6, 3.0)
    x = np.arange(len(mt))
    # Gradient-effect: stack two fills
    ax.fill_between(x, mt["Sales"], alpha=0.18, color=color, zorder=1)
    ax.fill_between(x, mt["Sales"] * 0.3, alpha=0.10, color=color, zorder=1)
    ax.plot(x, mt["Sales"], color=color, linewidth=2.2, zorder=3,
            marker="o", markersize=4, markerfacecolor="white",
            markeredgecolor=color, markeredgewidth=1.8)

    # Peak marker + annotation
    peak_i = mt["Sales"].idxmax()
    peak_v = mt.loc[peak_i, "Sales"]
    ax.plot(peak_i, peak_v, marker="*", markersize=13, color="#FBBF24", zorder=5)
    ax.annotate(f"Peak\n{_fmt(peak_v)}", xy=(peak_i, peak_v),
                xytext=(peak_i + 0.6, peak_v * 0.97),
                fontsize=6.5, color="#D97706",
                arrowprops=dict(arrowstyle="-", color="#D97706", lw=0.8))

    # MoM growth bar underneath (secondary axis)
    ax2 = ax.twinx()
    colors_mom = ["#16A34A" if v >= 0 else "#DC2626" for v in mt["MoM"]]
    ax2.bar(x, mt["MoM"] * 100, color=colors_mom, alpha=0.18, width=0.7, zorder=0)
    ax2.set_ylabel("MoM %", fontsize=6.5, color="#9CA3AF")
    ax2.tick_params(labelsize=6, colors="#9CA3AF")
    ax2.spines["top"].set_visible(False)
    ax2.spines["right"].set_color("#E5E7EB")
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}%"))

    _chart_title(ax, f"Monthly {sc_lbl} Trend", "#065F46",
                 "Bars = MoM growth  |  ★ = peak month")
    lbls = [str(d)[:7] for d in mt["Date"]]
    step = max(1, len(lbls) // 7)
    ax.set_xticks(x[::step])
    ax.set_xticklabels([lbls[i] for i in range(0, len(lbls), step)], rotation=30, fontsize=6.5)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: _fmt(v)))
    _clean_ax(ax, "y")
    fig.tight_layout(pad=0.7)
    return _save(fig)


# ── 4. Sales vs Profit scatter — density shading + regression band ───────────
def chart_scatter(df, ct):
    sc, pc = _sales_col(df, ct), _profit_col(df, ct)
    if not sc or not pc: return None
    samp = df[[sc, pc]].dropna().sample(min(600, len(df)), random_state=42)
    color = CHART_PALETTES["scatter"]

    fig, ax = _base_fig(5.6, 3.0)
    ax.scatter(samp[sc], samp[pc], alpha=0.35, color=color, s=14,
               edgecolors="white", linewidths=0.3, zorder=3)

    # Regression line + confidence band
    try:
        z = np.polyfit(samp[sc], samp[pc], 1)
        p_fn = np.poly1d(z)
        xs = np.linspace(samp[sc].min(), samp[sc].max(), 200)
        ys = p_fn(xs)
        # Simple ±std band
        residuals = samp[pc] - p_fn(samp[sc])
        std = residuals.std()
        ax.plot(xs, ys, color=color, lw=1.8, alpha=0.85, linestyle="--", zorder=4)
        ax.fill_between(xs, ys - std, ys + std, alpha=0.08, color=color, zorder=2)
        # Slope annotation
        slope_str = f"slope = {z[0]:.2f}"
        ax.text(0.97, 0.05, slope_str, transform=ax.transAxes,
                ha="right", fontsize=7, color=color,
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                          edgecolor=color, alpha=0.8))
    except Exception:
        pass

    # Zero-profit line
    if samp[pc].min() < 0 < samp[pc].max():
        ax.axhline(0, color="#DC2626", linewidth=0.7, linestyle=":", alpha=0.6)

    _chart_title(ax, f"{sc} vs {pc}", "#C2410C",
                 "Shaded band = ±1 std dev  |  dashed = trend")
    ax.set_xlabel(sc, fontsize=7, color="#6B7280")
    ax.set_ylabel(pc, fontsize=7, color="#6B7280")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: _fmt(v)))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: _fmt(v)))
    _clean_ax(ax, "y")
    fig.tight_layout(pad=0.7)
    return _save(fig)


# ── 5. Monthly order volume — bars + rolling avg + record count label ────────
def chart_orders(df, ct):
    dc = _date_col(df, ct)
    if not dc: return None
    color = CHART_PALETTES["orders"]
    tmp = df.copy()
    tmp[dc] = pd.to_datetime(tmp[dc], errors="coerce")
    tmp["_p"] = tmp[dc].dt.to_period("M").dt.to_timestamp()
    counts = tmp.groupby("_p").size().reset_index(name="Orders")
    if counts.empty: return None

    fig, ax = _base_fig(5.6, 3.0)
    x = np.arange(len(counts))

    # Colour bars by above/below mean
    mean_v = counts["Orders"].mean()
    bar_colors = [color if v >= mean_v else "#93C5FD" for v in counts["Orders"]]
    bars = ax.bar(x, counts["Orders"], color=bar_colors, alpha=0.85,
                  edgecolor="white", linewidth=0.5, width=0.72, zorder=3)

    # Mean reference line
    ax.axhline(mean_v, color="#6B7280", linewidth=0.9, linestyle=":",
               alpha=0.7, zorder=2, label=f"Avg {mean_v:.0f}")

    # 3-month rolling average
    if len(counts) >= 3:
        roll = counts["Orders"].rolling(3, center=True, min_periods=1).mean()
        ax.plot(x, roll, color="#FBBF24", linewidth=2.2, zorder=5,
                marker="D", markersize=3.5, markerfacecolor="#FBBF24",
                markeredgecolor="white", markeredgewidth=0.8, label="3-mo avg")

    ax.legend(fontsize=6.5, frameon=False, loc="upper left")
    _chart_title(ax, "Monthly Order Volume", "#1D4ED8",
                 f"Total records: {len(df):,}")
    lbls = [str(d)[:7] for d in counts["_p"]]
    step = max(1, len(lbls) // 7)
    ax.set_xticks(x[::step])
    ax.set_xticklabels([lbls[i] for i in range(0, len(lbls), step)], rotation=30, fontsize=6.5)
    _clean_ax(ax, "y")
    fig.tight_layout(pad=0.7)
    return _save(fig)


# ── 6. Seasonality heatmap — annotated cells + peak highlight ────────────────
def chart_heatmap(df, ct):
    dc, sc = _date_col(df, ct), _sales_col(df, ct)
    if not dc or not sc: return None
    tmp = df.copy()
    tmp[dc] = pd.to_datetime(tmp[dc], errors="coerce")
    tmp = tmp.dropna(subset=[dc])
    if tmp.empty: return None
    tmp["month"] = tmp[dc].dt.month.astype(int)
    tmp["year"]  = tmp[dc].dt.year.astype(int)
    piv = tmp.pivot_table(index="year", columns="month", values=sc, aggfunc="sum").fillna(0)
    piv.columns = piv.columns.astype(int)
    piv.index   = piv.index.astype(int)

    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    fig, ax = _base_fig(5.6, max(2.4, piv.shape[0] * 0.6 + 0.8))
    im = ax.imshow(piv.values, cmap=CHART_PALETTES["heatmap"],
                   aspect="auto", interpolation="nearest")

    ax.set_xticks(range(piv.shape[1]))
    ax.set_xticklabels([months[m-1] for m in piv.columns], fontsize=7.5)
    ax.set_yticks(range(piv.shape[0]))
    ax.set_yticklabels([str(y) for y in piv.index], fontsize=7.5)

    # Annotate each cell with abbreviated value
    for r in range(piv.shape[0]):
        for c in range(piv.shape[1]):
            v = piv.values[r, c]
            if v > 0:
                txt_c = "white" if v > piv.values.max() * 0.6 else "#374151"
                ax.text(c, r, _fmt(v).replace("$",""), ha="center", va="center",
                        fontsize=5.5, color=txt_c, fontweight="bold")

    # Gold border on peak cell
    flat_max = np.unravel_index(np.argmax(piv.values), piv.values.shape)
    ax.add_patch(plt.Rectangle(
        (flat_max[1]-0.5, flat_max[0]-0.5), 1, 1,
        fill=False, edgecolor="#FBBF24", lw=2.5, zorder=5))

    cbar = fig.colorbar(im, ax=ax, fraction=0.025, pad=0.03)
    cbar.ax.tick_params(labelsize=6)
    cbar.set_label("Sales", fontsize=6.5, color="#6B7280")
    _chart_title(ax, f"{sc} Seasonality Heatmap", "#6D28D9",
                 "Gold border = peak cell")
    ax.spines[:].set_visible(False)
    ax.tick_params(length=0)
    fig.tight_layout(pad=0.7)
    return _save(fig)


# ════════════════════════════════════════════════════════════════════════════════
# REPORTLAB PDF HELPERS
# ════════════════════════════════════════════════════════════════════════════════
def _ps(name, **kw):
    return ParagraphStyle(name, **kw)

def _styles():
    return {
        "section": _ps("sec", fontName="Helvetica-Bold", fontSize=11.5,
                       textColor=colors.HexColor(BRAND["deep"]),
                       spaceBefore=5, spaceAfter=2),
        "sub":     _ps("sub", fontName="Helvetica-Bold", fontSize=9,
                       textColor=colors.HexColor(BRAND["mid"]),
                       spaceBefore=4, spaceAfter=2),
        "body":    _ps("body", fontName="Helvetica", fontSize=8.5,
                       textColor=colors.HexColor(BRAND["ink"]),
                       leading=13, spaceAfter=2),
        "bullet":  _ps("blt", fontName="Helvetica", fontSize=8.5,
                       textColor=colors.HexColor(BRAND["ink"]),
                       leading=13, spaceAfter=2, leftIndent=12),
        "note":    _ps("note", fontName="Helvetica-Oblique", fontSize=7,
                       textColor=colors.HexColor(BRAND["muted"])),
    }

def _section_block(text, ST):
    return [
        Paragraph(text, ST["section"]),
        HRFlowable(width="100%", thickness=0.9,
                   color=colors.HexColor(BRAND["soft"]), spaceAfter=5),
    ]

def _hf(canvas, doc):
    canvas.saveState()
    W, H = A4

    # Header band
    canvas.setFillColor(colors.HexColor(BRAND["deep"]))
    canvas.rect(0, H - 20*mm, W, 20*mm, fill=1, stroke=0)
    # Teal accent strip
    canvas.setFillColor(colors.HexColor(BRAND["teal"]))
    canvas.rect(0, H - 20*mm, 4*mm, 20*mm, fill=1, stroke=0)
    # Title
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 11.5)
    canvas.drawString(10*mm, H - 12*mm, "Zero Click AI  ·  Comprehensive Analysis Report")
    # Sub-line
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.HexColor(BRAND["soft"]))
    canvas.drawString(10*mm, H - 17.5*mm,
                      f"Report Date: {NOW.strftime('%d %b %Y')}  |  "
                      f"Generated: {NOW.strftime('%H:%M')}  |  Genesis Training")
    canvas.drawRightString(W - 10*mm, H - 17.5*mm, "Source: Zero Click AI")

    # Footer
    canvas.setFillColor(colors.HexColor(BRAND["pale"]))
    canvas.rect(0, 0, W, 10*mm, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor(BRAND["teal"]))
    canvas.rect(0, 0, 3*mm, 10*mm, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor(BRAND["muted"]))
    canvas.setFont("Helvetica", 7)
    canvas.drawCentredString(W/2, 3.5*mm,
                             f"Page {doc.page}  |  Confidential — For Internal Use Only")
    canvas.restoreState()

def _kpi_row(kpi_triples, start_idx, pw):
    col_w = (pw - 8*mm) / 3
    cells = []
    for i, (lbl, val, dlt) in enumerate(kpi_triples):
        idx     = (start_idx + i) % len(KPI_CARDS)
        bg, acc = KPI_CARDS[idx]
        is_pos  = not str(dlt).lstrip().startswith("-")
        dlt_col = colors.HexColor(BRAND["green"]) if is_pos else colors.HexColor(BRAND["red"])
        sym     = "▲" if is_pos else "▼"
        inner = Table(
            [[Paragraph(str(val)[:14],
                        _ps(f"kv{i}", fontName="Helvetica-Bold", fontSize=15,
                            textColor=colors.HexColor(acc), alignment=TA_CENTER))],
             [Paragraph(f"{sym}  {dlt}",
                        _ps(f"kd{i}", fontName="Helvetica", fontSize=7.5,
                            textColor=dlt_col, alignment=TA_CENTER))],
             [Paragraph(str(lbl)[:26],
                        _ps(f"kl{i}", fontName="Helvetica-Bold", fontSize=7,
                            textColor=colors.HexColor(BRAND["muted"]),
                            alignment=TA_CENTER))]],
            colWidths=[col_w],
            style=TableStyle([
                ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor(bg)),
                ("BOX",           (0,0),(-1,-1), 1.2, colors.HexColor(acc)),
                ("LINEABOVE",     (0,0),(-1,0),  3,   colors.HexColor(acc)),
                ("TOPPADDING",    (0,0),(-1,-1), 8),
                ("BOTTOMPADDING", (0,0),(-1,-1), 8),
                ("ALIGN",         (0,0),(-1,-1), "CENTER"),
                ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ])
        )
        cells.append(inner)
    return Table([cells],
                 colWidths=[col_w + 2.5*mm]*3,
                 style=TableStyle([
                     ("ALIGN",         (0,0),(-1,-1), "CENTER"),
                     ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
                     ("LEFTPADDING",   (0,0),(-1,-1), 2),
                     ("RIGHTPADDING",  (0,0),(-1,-1), 2),
                 ]))

def _data_table(headers, rows, col_w_mm):
    cw   = [w * mm for w in col_w_mm]
    data = [headers] + rows
    style = [
        ("BACKGROUND",    (0,0),(-1,0),  colors.HexColor(BRAND["deep"])),
        ("TEXTCOLOR",     (0,0),(-1,0),  colors.white),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0),  8),
        ("FONTNAME",      (0,1),(-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1),(-1,-1), 7.5),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor(BRAND["rule"])),
        ("TOPPADDING",    (0,0),(-1,-1), 3.5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3.5),
        ("LINEBELOW",     (0,0),(-1,0),  2, colors.HexColor(BRAND["teal"])),
    ]
    for r in range(1, len(data)):
        bg = BRAND["row_a"] if r % 2 == 1 else BRAND["row_b"]
        style.append(("BACKGROUND", (0,r),(-1,r), colors.HexColor(bg)))
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle(style))
    return tbl

def _chart_img(buf, w_mm, h_mm):
    if buf is None: return Spacer(w_mm*mm, h_mm*mm)
    return RLImage(buf, width=w_mm*mm, height=h_mm*mm)


# ════════════════════════════════════════════════════════════════════════════════
# MAIN PDF GENERATOR
# ════════════════════════════════════════════════════════════════════════════════
def generate_report_pdf(df, kpis, ml_results, forecast_data, insights, charts) -> bytes:
    """ml_results accepted for API compatibility but ignored."""
    if not HAS_RL:
        return b""

    ct = _col_types(df)
    ST = _styles()
    pw = A4[0] - 24*mm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=24*mm, bottomMargin=14*mm,
        leftMargin=12*mm, rightMargin=12*mm,
    )
    story = []

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 1 — EXECUTIVE DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    banner = Table(
        [[Paragraph("EXECUTIVE PERFORMANCE DASHBOARD",
                    _ps("bh", fontName="Helvetica-Bold", fontSize=12,
                        textColor=colors.white, alignment=TA_CENTER))]],
        colWidths=[pw],
        style=TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor(BRAND["mid"])),
            ("TOPPADDING",    (0,0),(-1,-1), 9),
            ("BOTTOMPADDING", (0,0),(-1,-1), 9),
            ("LINEABOVE",     (0,0),(-1,0),  3, colors.HexColor(BRAND["teal"])),
            ("LINEBELOW",     (0,0),(-1,-1), 3, colors.HexColor(BRAND["teal"])),
        ])
    )
    story.append(banner)
    story.append(Spacer(1, 5*mm))

    kpi_list = list(kpis or [])
    while len(kpi_list) < 6:
        kpi_list.append(("N/A", "—", "—"))

    def _ktriple(k):
        if len(k) >= 3: return (str(k[0]), str(k[1]), str(k[2]))
        if len(k) == 2: return (str(k[0]), str(k[1]), "—")
        return (str(k[0]), "—", "—")

    story.append(_kpi_row([_ktriple(kpi_list[i]) for i in range(3)], 0, pw))
    story.append(Spacer(1, 4*mm))
    story.append(_kpi_row([_ktriple(kpi_list[i]) for i in range(3, 6)], 3, pw))
    story.append(Spacer(1, 7*mm))

    # AI Summary
    ai_hdr = Table(
        [[Paragraph("  ◈  AI NARRATED SUMMARY",
                    _ps("aih", fontName="Helvetica-Bold", fontSize=9,
                        textColor=colors.HexColor(BRAND["deep"])))]],
        colWidths=[pw],
        style=TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), colors.HexColor(BRAND["pale"])),
            ("LINEABOVE",     (0,0),(-1,0),  3, colors.HexColor(BRAND["teal"])),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
        ])
    )
    story.append(ai_hdr)

    ai_short = ""
    if isinstance(insights, str):
        lines = [l.strip() for l in insights.split("\n")
                 if l.strip() and not l.strip().startswith("#")]
        ai_short = " ".join(lines[:5])[:700]
    elif isinstance(insights, list):
        ai_short = " ".join(str(i) for i in insights[:5])[:700]

    ai_body = Table(
        [[Paragraph(ai_short or "Analysis complete.", ST["body"])]],
        colWidths=[pw],
        style=TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), colors.white),
            ("BOX",           (0,0),(-1,-1), 0.6, colors.HexColor(BRAND["rule"])),
            ("TOPPADDING",    (0,0),(-1,-1), 9),
            ("BOTTOMPADDING", (0,0),(-1,-1), 9),
            ("LEFTPADDING",   (0,0),(-1,-1), 12),
            ("RIGHTPADDING",  (0,0),(-1,-1), 12),
        ])
    )
    story.append(ai_body)
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "*Profit margin = Profit / Sales  |  Data reflects the period shown in available records.",
        ST["note"]))

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 2 — VISUAL ANALYTICS
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.extend(_section_block("Visual Analytics", ST))

    chart_defs = [
        (f"{_sales_col(df,ct) or 'Sales'} by Region",
         lambda: chart_region(df, ct)),
        (f"Sales by {_cat_col(df,ct) or 'Category'}",
         lambda: chart_category(df, ct)),
        (f"Monthly {_sales_col(df,ct) or 'Sales'} Trend",
         lambda: chart_trend(df, ct)),
        (f"{_sales_col(df,ct) or 'Sales'} vs {_profit_col(df,ct) or 'Profit'}",
         lambda: chart_scatter(df, ct)),
        ("Monthly Order Volume",
         lambda: chart_orders(df, ct)),
        (f"{_sales_col(df,ct) or 'Sales'} Seasonality",
         lambda: chart_heatmap(df, ct)),
    ]

    rendered = []
    for title, fn in chart_defs:
        try:
            b = fn()
            if b: rendered.append((title, b))
        except Exception as e:
            print(f"  Chart '{title}' skipped: {e}")

    CW = (pw - 5*mm) / 2
    CH = 68   # mm — slightly taller for richer charts

    for i in range(0, len(rendered), 2):
        lt, lb = rendered[i]
        rt, rb = rendered[i+1] if i+1 < len(rendered) else (None, None)
        li = _chart_img(lb, CW/mm, CH)
        ri = _chart_img(rb, CW/mm, CH) if rb else Spacer(CW, CH*mm)

        def _cap(t):
            return Paragraph(t or "",
                             _ps("ct", fontName="Helvetica-Bold", fontSize=7.5,
                                 textColor=colors.HexColor(BRAND["muted"]),
                                 alignment=TA_CENTER))

        grid = Table(
            [[_cap(lt), _cap(rt)], [li, ri]],
            colWidths=[CW, CW],
            style=TableStyle([
                ("ALIGN",         (0,0),(-1,-1), "CENTER"),
                ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
                ("LEFTPADDING",   (0,0),(-1,-1), 2),
                ("RIGHTPADDING",  (0,0),(-1,-1), 2),
                ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ])
        )
        story.append(grid)
        if i + 2 < len(rendered):
            story.append(Spacer(1, 2*mm))

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 3 — TOP ITEMS + AI INSIGHTS (two-column)
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.extend(_section_block("Top Records by Key Metric", ST))

    tp   = _top_items(df, ct, n=16)
    hdrs = list(tp.columns)
    nc   = len(hdrs)
    t_pw = pw * 0.55
    i_pw = pw * 0.41

    col_w = [8] + [int((t_pw/mm - 10) / (nc-1))] * (nc-1)
    t_rows = [[str(tp.iloc[r][c])[:20] for c in tp.columns] for r in range(len(tp))]
    prod_tbl = _data_table(hdrs, t_rows, col_w)

    # AI insights sidebar
    full_ins = (insights if isinstance(insights, str)
                else "\n".join(f"• {i}" for i in (insights or [])))
    ins_items = [
        Paragraph("AI Insights", ST["section"]),
        HRFlowable(width="100%", thickness=0.9,
                   color=colors.HexColor(BRAND["soft"]), spaceAfter=5),
    ]
    for line in full_ins.split("\n"):
        line = line.strip()
        if not line:
            ins_items.append(Spacer(1, 2*mm))
        elif line.startswith("## "):
            ins_items.append(Paragraph(line[3:], ST["sub"]))
        elif line.startswith(("- ","• ","* ")):
            ins_items.append(Paragraph(f"• {line[2:]}", ST["bullet"]))
        else:
            ins_items.append(Paragraph(line, ST["body"]))

    ins_tbl = Table(
        [[item] for item in ins_items],
        colWidths=[i_pw],
        style=TableStyle([
            ("TOPPADDING",    (0,0),(-1,-1), 1),
            ("BOTTOMPADDING", (0,0),(-1,-1), 1),
            ("LEFTPADDING",   (0,0),(-1,-1), 8),
            ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ])
    )

    two_col = Table(
        [[prod_tbl, ins_tbl]],
        colWidths=[t_pw + 2*mm, i_pw],
        style=TableStyle([
            ("VALIGN",       (0,0),(-1,-1), "TOP"),
            ("LEFTPADDING",  (0,0),(-1,-1), 0),
            ("RIGHTPADDING", (0,0),(-1,-1), 0),
            ("LINEBEFORE",   (1,0),(1,-1), 0.5, colors.HexColor(BRAND["rule"])),
        ])
    )
    story.append(two_col)

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 4 — DATASET SAMPLE
    # ══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.extend(_section_block("Dataset Sample — First 20 Rows", ST))

    sample = df.head(20).astype(str)
    cols   = list(sample.columns[:8])
    nc2    = len(cols)
    cw2    = [int(pw/mm / nc2)] * nc2
    s_rows = [[str(sample.iloc[r][c])[:16] for c in cols] for r in range(len(sample))]
    story.append(_data_table([c[:14] for c in cols], s_rows, cw2))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"Showing first 20 of {len(df):,} records  |  {len(df.columns)} columns total.",
        ST["note"]))

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 5 (OPTIONAL) — FORECAST
    # ══════════════════════════════════════════════════════════════════════════
    if forecast_data:
        story.append(PageBreak())
        story.extend(_section_block("Time-Series Forecast", ST))
        story.append(Paragraph(
            f"Model: {forecast_data.get('model','N/A')}  |  Horizon: 6 months", ST["body"]))
        story.append(Spacer(1, 3*mm))
        fdf = forecast_data.get("forecast_df")
        if fdf is not None and not fdf.empty:
            f_hdrs = ["Period", "Predicted", "Lower 95%", "Upper 95%"]
            f_cw   = [50, 44, 44, 44]
            f_rows = []
            for _, row in fdf.head(12).iterrows():
                dt = str(row.get("ds","")).split(" ")[0]
                f_rows.append([dt,
                               f"${_num(row.get('yhat',0)):,.0f}",
                               f"${_num(row.get('yhat_lower',0)):,.0f}",
                               f"${_num(row.get('yhat_upper',0)):,.0f}"])
            story.append(_data_table(f_hdrs, f_rows, f_cw))

    doc.build(story, onFirstPage=_hf, onLaterPages=_hf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT — polished, conditional formatting, data bars
# ════════════════════════════════════════════════════════════════════════════════
def _xl_hdr(ws, row, headers, accent):
    fill = PatternFill("solid", fgColor=accent)
    fnt  = Font(bold=True, color="FFFFFF", size=10)
    bdr  = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="medium"))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = fill; c.font = fnt; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(h))+5)
    ws.row_dimensions[row].height = 22

def _xl_data_row(ws, ri, vals, alt=False):
    hex_bg = "F5F3FF" if alt else "FFFFFF"
    fill = PatternFill("solid", fgColor=hex_bg)
    bdr  = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = fill; c.border = bdr
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")

def _xl_title(ws, text, row=1, end_col=8, accent="4C1D95"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=accent)
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

def _xl_subtitle(ws, text, row, end_col=8, accent="0D9488"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=accent)
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18

def _write_df(ws, df_in, start=3, accent="4C1D95"):
    hdrs = list(df_in.columns)
    _xl_hdr(ws, start, hdrs, accent)
    for ri, (_, row) in enumerate(df_in.iterrows(), 1):
        vals = list(row)
        _xl_data_row(ws, start+ri, vals, alt=ri%2==0)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=start+ri, column=ci)
            cn   = hdrs[ci-1].lower()
            if any(kw in cn for kw in ["mom","growth","pct","share"]):
                try:
                    v = float(val)
                    cell.font = Font(size=9, bold=True,
                                     color="166534" if v>=0 else "991B1B")
                    cell.value = f"+{v*100:.2f}%" if v>=0 else f"{v*100:.2f}%"
                except: pass
            elif any(kw in cn for kw in ["sales","revenue","profit","amount","price","total","fee"]):
                try: cell.value = float(val); cell.number_format = '#,##0.00'
                except: pass

def _apply_databar(ws, col_letter, min_row, max_row, color="638EC6"):
    """Apply a data bar conditional format to a column range."""
    try:
        col_range = f"{col_letter}{min_row}:{col_letter}{max_row}"
        rule = DataBarRule(start_type="min", start_value=0,
                           end_type="max", end_value=100,
                           color=color)
        ws.conditional_formatting.add(col_range, rule)
    except Exception:
        pass

def _apply_colorscale(ws, col_letter, min_row, max_row):
    """Green-yellow-red color scale."""
    try:
        col_range = f"{col_letter}{min_row}:{col_letter}{max_row}"
        rule = ColorScaleRule(
            start_type="min",  start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",    end_color="63BE7B",
        )
        ws.conditional_formatting.add(col_range, rule)
    except Exception:
        pass

def _parse_kpi_numeric(s):
    s = str(s).strip().replace(",","").replace("$","").replace("%","")
    s = s.replace("B","e9").replace("M","e6").replace("K","e3").replace("k","e3")
    try:    return float(s)
    except: return None


def generate_report_excel(df, kpis, ml_results, forecast_data, insights) -> bytes:
    """ml_results accepted for API compatibility but ignored."""
    if not HAS_OXL:
        return b""

    ct = _col_types(df)
    wb = openpyxl.Workbook()
    sc_lbl = _sales_col(df, ct) or "Sales"
    rg_lbl = _reg_col(df, ct)  or "Region"
    ct_lbl = _cat_col(df, ct)  or "Category"

    # ── README ────────────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "README"
    acc = XL_ACCENTS["README"]
    _xl_title(ws, "Zero Click AI  ·  Comprehensive Analysis Report", 1, 6, acc)
    _xl_subtitle(ws, "Report Summary", 2, 6, XL_ACCENTS["KPIs"])
    info = [
        ("Report Date",   NOW.strftime("%d %b %Y  %H:%M")),
        ("Platform",      "Zero Click AI  ·  Genesis Training"),
        ("Total Rows",    f"{len(df):,}"),
        ("Total Columns", f"{len(df.columns)}"),
        ("AI Summary",    (insights[:400] if isinstance(insights,str) else str(insights)[:400])),
    ]
    for ri, (k, v) in enumerate(info, 3):
        ws.cell(row=ri, column=1, value=k).font = Font(bold=True, color=acc, size=10)
        c2 = ws.cell(row=ri, column=2, value=v)
        c2.font = Font(size=10)
        c2.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[ri].height = 18
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 75

    # ── KPIs ──────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("KPIs")
    acc2 = XL_ACCENTS["KPIs"]
    _xl_title(ws2, "Key Performance Indicators", 1, 5, XL_ACCENTS["README"])
    _xl_subtitle(ws2, "Core business metrics at a glance", 2, 5, acc2)
    kpi_rows = [[str(k[0]), str(k[1]), str(k[2]) if len(k)>2 else ""]
                for k in (kpis or [])]
    _xl_hdr(ws2, 3, ["Metric", "Value", "Description", "", "Numeric Value"], acc2)
    for i, r in enumerate(kpi_rows):
        _xl_data_row(ws2, i+4, r + ["",""], alt=i%2==0)
        nv = _parse_kpi_numeric(r[1]) if len(r)>1 else None
        if nv is not None:
            cell = ws2.cell(row=i+4, column=5, value=nv)
            cell.font = Font(size=9, bold=True, color="047857")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor="ECFDF5")
            cell.number_format = '#,##0.00'
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 42
    ws2.column_dimensions["E"].width = 22
    n_kpi = len(kpi_rows)
    # Apply data bar to numeric column
    if n_kpi >= 2:
        _apply_databar(ws2, "E", 4, 3+n_kpi, color=acc2)
        ch = BarChart(); ch.type = "bar"; ch.title = "KPI Overview"
        ch.style = 26; ch.width = 22; ch.height = max(10, n_kpi*1.5)
        ch.grouping = "clustered"
        data = Reference(ws2, min_col=5, min_row=3, max_row=3+n_kpi)
        cats = Reference(ws2, min_col=1, min_row=4, max_row=3+n_kpi)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ch.series[0].graphicalProperties.solidFill = acc2
        ch.plot_area.layout = None
        ws2.add_chart(ch, "G3")

    # ── Sales by Region ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Sales_by_Region")
    acc3 = XL_ACCENTS["Sales_by_Region"]
    _xl_title(ws3, f"{sc_lbl} by {rg_lbl}", 1, 4, XL_ACCENTS["README"])
    _xl_subtitle(ws3, "Regional performance — sorted descending", 2, 4, acc3)
    rbr = _sales_by_region(df, ct).sort_values("Sales", ascending=False)
    _write_df(ws3, rbr, start=3, accent=acc3)
    n_rbr = len(rbr)
    # Data bar on Sales column
    if n_rbr >= 1:
        _apply_databar(ws3, "B", 4, 3+n_rbr, color=acc3)
    if n_rbr >= 2:
        ch = BarChart(); ch.title = f"{sc_lbl} by {rg_lbl}"
        ch.style = 26; ch.width = 18; ch.height = 12; ch.type = "bar"
        data = Reference(ws3, min_col=2, min_row=3, max_row=3+n_rbr)
        cats = Reference(ws3, min_col=1, min_row=4, max_row=3+n_rbr)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ch.series[0].graphicalProperties.solidFill = acc3
        ws3.add_chart(ch, "D3")

    # ── Sales by Category ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Sales_by_Category")
    acc4 = XL_ACCENTS["Sales_by_Category"]
    _xl_title(ws4, f"{sc_lbl} by {ct_lbl}", 1, 5, XL_ACCENTS["README"])
    _xl_subtitle(ws4, "Category breakdown with share %", 2, 5, acc4)
    sbc = _sales_by_category(df, ct).sort_values("Sales", ascending=False)
    _write_df(ws4, sbc, start=3, accent=acc4)
    n_sbc = len(sbc)
    if n_sbc >= 1:
        _apply_databar(ws4, "B", 4, 3+n_sbc, color=acc4)
        _apply_colorscale(ws4, "C", 4, 3+n_sbc)   # color scale on Pct
    if n_sbc >= 2:
        ch = BarChart(); ch.title = f"{sc_lbl} by {ct_lbl}"
        ch.style = 26; ch.width = 20; ch.height = 14; ch.type = "bar"
        data = Reference(ws4, min_col=2, min_row=3, max_row=3+n_sbc)
        cats = Reference(ws4, min_col=1, min_row=4, max_row=3+n_sbc)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ch.series[0].graphicalProperties.solidFill = acc4
        ws4.add_chart(ch, "E3")

    # ── Top Products ──────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Top_Products")
    acc5 = XL_ACCENTS["Top_Products"]
    _xl_title(ws5, "Top Items by Revenue", 1, 6, XL_ACCENTS["README"])
    _xl_subtitle(ws5, "Ranked by total sales — profit cells colour-coded", 2, 6, acc5)
    tp = _top_items(df, ct, n=10)
    _write_df(ws5, tp, start=3, accent=acc5)
    # Colour-code profit column + data bar on Sales
    if "Profit" in tp.columns:
        pci = list(tp.columns).index("Profit") + 1
        for ri in range(4, 4+len(tp)):
            cell = ws5.cell(row=ri, column=pci)
            try:
                v = float(cell.value)
                cell.fill = PatternFill("solid", fgColor="DCFCE7" if v>0 else "FEE2E2")
                cell.font = Font(color="166534" if v>0 else "991B1B", size=9, bold=True)
                cell.number_format = '#,##0.00'; cell.value = v
            except: pass
    sci = list(tp.columns).index("Sales") + 1
    _apply_databar(ws5, get_column_letter(sci), 4, 3+len(tp), color=acc5)

    # ── Monthly Trend ─────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Monthly_Trend")
    acc6 = XL_ACCENTS["Monthly_Trend"]
    _xl_title(ws6, f"Monthly {sc_lbl} Trend", 1, 5, XL_ACCENTS["README"])
    _xl_subtitle(ws6, "Month-over-month performance  |  green = growth, red = decline", 2, 5, acc6)
    mt = _monthly_trend(df, ct)
    if mt.empty:
        ws6.cell(row=4, column=1, value="No datetime column detected.")
    else:
        _write_df(ws6, mt, start=3, accent=acc6)
        n_mt = len(mt)
        _apply_databar(ws6, "B", 4, 3+n_mt, color=acc6)
        _apply_colorscale(ws6, "C", 4, 3+n_mt)    # color scale on MoM
        ch = LineChart(); ch.title = f"Monthly {sc_lbl}"; ch.style = 10
        ch.width = 22; ch.height = 12
        data = Reference(ws6, min_col=2, min_row=3, max_row=3+n_mt)
        ch.add_data(data, titles_from_data=True)
        ch.series[0].graphicalProperties.solidFill = acc6
        ch.series[0].graphicalProperties.line.solidFill = acc6
        ch.series[0].graphicalProperties.line.width = 20000   # EMUs ~1.5pt
        ch.series[0].smooth = True
        ws6.add_chart(ch, "E3")

    # ── Raw Data ──────────────────────────────────────────────────────────────
    ws7 = wb.create_sheet("Raw_Data")
    acc7 = XL_ACCENTS["Raw_Data"]
    end  = min(len(df.columns), 14)
    _xl_title(ws7, "Raw Cleaned Dataset  ·  First 500 rows", 1, end, XL_ACCENTS["README"])
    _xl_subtitle(ws7, f"{len(df):,} total rows  ·  {len(df.columns)} columns", 2, end, acc7)
    _write_df(ws7, df.head(500), start=3, accent=acc7)

    # Freeze panes + auto-filter on all data sheets
    for sheet in [ws2, ws3, ws4, ws5, ws6, ws7]:
        sheet.freeze_panes = "A4"
    for sheet in [ws3, ws4, ws5, ws6, ws7]:
        sheet.auto_filter.ref = sheet.dimensions

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
